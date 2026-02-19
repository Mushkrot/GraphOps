"""Excel Parser — reads sheets according to IngestionSpec and produces staged data.

Parses Excel files using openpyxl, maps columns to entities and relationships,
and computes dual hashes (raw + normalized) for each row.
"""

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

import openpyxl

from backend.core.hashing import (
    compute_normalized_hash,
    compute_raw_hash,
)
from backend.core.ingestion_spec import (
    ColumnMapping,
    EntityMapping,
    IngestionSpec,
    RelationshipMapping,
    SheetSpec,
)

logger = logging.getLogger(__name__)


@dataclass
class StagedEntity:
    """An entity extracted from a row, ready for graph ingestion."""
    entity_type: str
    primary_key: str
    display_name: Optional[str]
    properties: dict[str, Any]
    source_ref: str


@dataclass
class StagedRelationship:
    """A relationship extracted from a row, ready for graph ingestion."""
    relationship_type: str
    from_entity_type: str
    from_primary_key: str
    to_entity_type: str
    to_primary_key: str
    properties: Optional[dict[str, Any]] = None
    source_ref: str = ""


@dataclass
class StagedRow:
    """A parsed row with staged entities, relationships, and computed hashes."""
    row_index: int
    raw_values: list[Any]
    entities: list[StagedEntity]
    relationships: list[StagedRelationship]
    raw_hash: str
    normalized_hash: str


def _build_header_map(headers: list[Any]) -> dict[str, int]:
    """Map column header names to their 0-based index."""
    header_map = {}
    for i, h in enumerate(headers):
        if h is not None:
            header_map[str(h).strip()] = i
    return header_map


def _extract_row_values(row_cells: tuple, num_cols: int) -> list[Any]:
    """Extract cell values from an openpyxl row."""
    values = []
    for i in range(num_cols):
        if i < len(row_cells):
            values.append(row_cells[i].value if hasattr(row_cells[i], 'value') else row_cells[i])
        else:
            values.append(None)
    return values


def _get_cell_value(row_values: list[Any], header_map: dict[str, int], column_name: str) -> Any:
    """Get a cell value by column header name."""
    idx = header_map.get(column_name)
    if idx is None:
        return None
    if idx < len(row_values):
        return row_values[idx]
    return None


def _resolve_key(key_template: str, key_columns: list[str], row_data: dict[str, Any]) -> Optional[str]:
    """Resolve a key template using extracted row data.

    key_template: e.g., "{item_code}" or "{location_id}_{suffix}"
    key_columns: columns that must be non-null for key to be valid
    row_data: column_name -> value mapping
    """
    # Check that all key columns have values
    for col in key_columns:
        val = row_data.get(col)
        if val is None or (isinstance(val, str) and val.strip() == ""):
            return None

    try:
        return key_template.format(**{k: str(v) for k, v in row_data.items() if v is not None})
    except (KeyError, IndexError):
        return None


def _extract_entity(
    entity_key: str,
    mapping: EntityMapping,
    row_values: list[Any],
    header_map: dict[str, int],
    sheet_name: str,
    row_index: int,
) -> Optional[StagedEntity]:
    """Extract a single entity from a row according to the entity mapping."""
    # Build row_data from properties → target_property -> value
    row_data = {}
    properties = {}
    for prop in mapping.properties:
        value = _get_cell_value(row_values, header_map, prop.source_column)
        if prop.transform:
            value = _apply_transform(value, prop.transform)
        row_data[prop.target_property] = value
        properties[prop.target_property] = value

    # Resolve primary key
    primary_key = _resolve_key(mapping.key_template, mapping.key_columns, row_data)
    if primary_key is None:
        return None

    # Use first non-key property as display_name, or primary_key
    display_name = None
    for prop in mapping.properties:
        if prop.target_property not in mapping.key_columns:
            val = properties.get(prop.target_property)
            if val is not None:
                display_name = str(val)
                break
    if display_name is None:
        display_name = primary_key

    return StagedEntity(
        entity_type=mapping.entity_type,
        primary_key=primary_key,
        display_name=display_name,
        properties=properties,
        source_ref=f"sheet:{sheet_name},row:{row_index}",
    )


def _extract_relationship(
    rel_mapping: RelationshipMapping,
    entities_by_key: dict[str, StagedEntity],
    row_values: list[Any],
    header_map: dict[str, int],
    sheet_name: str,
    row_index: int,
) -> Optional[StagedRelationship]:
    """Extract a relationship between two entities from a row."""
    from_entity = entities_by_key.get(rel_mapping.from_entity)
    to_entity = entities_by_key.get(rel_mapping.to_entity)

    if from_entity is None or to_entity is None:
        return None

    # Extract relationship properties if any
    rel_props = None
    if rel_mapping.properties:
        rel_props = {}
        for prop in rel_mapping.properties:
            value = _get_cell_value(row_values, header_map, prop.source_column)
            if prop.transform:
                value = _apply_transform(value, prop.transform)
            rel_props[prop.target_property] = value

    return StagedRelationship(
        relationship_type=rel_mapping.relationship_type,
        from_entity_type=from_entity.entity_type,
        from_primary_key=from_entity.primary_key,
        to_entity_type=to_entity.entity_type,
        to_primary_key=to_entity.primary_key,
        properties=rel_props,
        source_ref=f"sheet:{sheet_name},row:{row_index}",
    )


def _apply_transform(value: Any, transform: str) -> Any:
    """Apply a simple transform to a value."""
    if value is None:
        return None
    s = str(value)
    if transform == "strip":
        return s.strip()
    if transform == "lower":
        return s.lower()
    if transform == "upper":
        return s.upper()
    if transform == "int":
        try:
            return int(float(s))
        except (ValueError, TypeError):
            return value
    if transform == "float":
        try:
            return float(s)
        except (ValueError, TypeError):
            return value
    return value


def _get_value_types_from_spec(spec: IngestionSpec, header_map: dict[str, int]) -> list[str]:
    """Build a list of value_types matching the column order for hash computation."""
    # For simplicity, all types default to "string"
    # A more sophisticated version could look at schema property types
    return ["string"] * len(header_map)


def parse_sheet(
    ws,
    sheet_spec: SheetSpec,
    spec: IngestionSpec,
    sheet_name: str,
) -> list[StagedRow]:
    """Parse a single worksheet according to its SheetSpec."""
    rows = list(ws.iter_rows())
    if not rows:
        return []

    # Extract headers
    header_row_idx = sheet_spec.header_row
    if header_row_idx >= len(rows):
        logger.warning(f"Header row {header_row_idx} out of range for sheet {sheet_name}")
        return []

    headers = [cell.value for cell in rows[header_row_idx]]
    header_map = _build_header_map(headers)
    num_cols = len(headers)

    skip_rows = set(sheet_spec.skip_rows or [])
    skip_rows.add(header_row_idx)

    value_types = ["string"] * num_cols

    staged_rows = []
    for row_idx, row in enumerate(rows):
        if row_idx in skip_rows:
            continue

        raw_values = [cell.value for cell in row[:num_cols]]

        # Skip completely empty rows
        if all(v is None for v in raw_values):
            continue

        # Compute hashes
        raw_hash = compute_raw_hash(raw_values, spec.raw_hash_serialization)
        normalized_hash = compute_normalized_hash(
            raw_values,
            spec.raw_hash_serialization,
            spec.change_detection.normalization_rules,
            value_types,
        )

        # Extract entities
        entities_by_key: dict[str, StagedEntity] = {}
        entity_list: list[StagedEntity] = []
        for entity_key, entity_mapping in sheet_spec.entities.items():
            entity = _extract_entity(
                entity_key, entity_mapping, raw_values, header_map,
                sheet_name, row_idx,
            )
            if entity:
                entities_by_key[entity_key] = entity
                entity_list.append(entity)

        # Extract relationships
        rel_list: list[StagedRelationship] = []
        for rel_mapping in sheet_spec.relationships:
            rel = _extract_relationship(
                rel_mapping, entities_by_key, raw_values, header_map,
                sheet_name, row_idx,
            )
            if rel:
                rel_list.append(rel)

        if entity_list:  # Only stage rows that produced at least one entity
            staged_rows.append(StagedRow(
                row_index=row_idx,
                raw_values=raw_values,
                entities=entity_list,
                relationships=rel_list,
                raw_hash=raw_hash,
                normalized_hash=normalized_hash,
            ))

    return staged_rows


def parse_excel(
    file_path: Path,
    spec: IngestionSpec,
) -> list[StagedRow]:
    """Parse an Excel file according to the ingestion spec.

    Returns a list of StagedRow objects with entities, relationships, and hashes.
    """
    wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)

    all_rows: list[StagedRow] = []
    for sheet_spec in spec.sheets:
        # Select sheet by name or index
        if sheet_spec.sheet_name:
            if sheet_spec.sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{sheet_spec.sheet_name}' not found in workbook")
                continue
            ws = wb[sheet_spec.sheet_name]
            sheet_name = sheet_spec.sheet_name
        elif sheet_spec.sheet_index is not None:
            if sheet_spec.sheet_index >= len(wb.sheetnames):
                logger.warning(f"Sheet index {sheet_spec.sheet_index} out of range")
                continue
            ws = wb.worksheets[sheet_spec.sheet_index]
            sheet_name = wb.sheetnames[sheet_spec.sheet_index]
        else:
            ws = wb.active
            sheet_name = ws.title if ws else "Sheet1"

        rows = parse_sheet(ws, sheet_spec, spec, sheet_name)
        all_rows.extend(rows)

    wb.close()
    return all_rows
